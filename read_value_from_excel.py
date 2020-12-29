from openpyxl import load_workbook
from collections import OrderedDict


def get_value_from_excel():
    wb = load_workbook(r"/tmp/case_template.xlsx")
    sheet_name = "case"
    target_sheet = wb.get_sheet_by_name(sheet_name)
    # openpyxl read excel index from row 1 and column 1
    case_dict = {}
    # insert info into a nested dict
    for i in range(2, target_sheet.max_row + 1):
        case_dict["case{}".format(i-1)] = OrderedDict()
        case_dict["case{}".format(i-1)]["aaa"] = target_sheet.cell(row=i, column=1).value
        case_dict["case{}".format(i-1)]["bbb"] = target_sheet.cell(row=i, column=2).value
        case_dict["case{}".format(i-1)]["ccc"] = target_sheet.cell(row=i, column=3).value
        case_dict["case{}".format(i-1)]["ddd"] = target_sheet.cell(row=i, column=4).value
        case_dict["case{}".format(i-1)]["eee"] = target_sheet.cell(row=i, column=5).value
        case_dict["case{}".format(i-1)]["fff"] = target_sheet.cell(row=i, column=6).value
        case_dict["case{}".format(i-1)]["ggg"] = target_sheet.cell(row=i, column=7).value
        case_dict["case{}".format(i-1)]["hhh"] = target_sheet.cell(row=i, column=8).value
    return case_dict


# print(get_value_from_excel())
