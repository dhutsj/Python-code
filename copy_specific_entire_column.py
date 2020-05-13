import openpyxl
from openpyxl.styles import Font


def copy_column():
    wb1 = openpyxl.load_workbook('aaa.xlsx')
    ws1 = wb1.active

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active

    for cell in ws1['I:I']:
        ws2.cell(row=cell.row, column=1, value=cell.value)

    for cell in ws1['O:O']:
        ws2.cell(row=cell.row, column=2, value=cell.value)

    row1 = ws2[1]
    for item in row1:
        item.font = Font(bold=True, size=11)

    wb2.save('Output.xlsx')


if __name__ == '__main__':
    copy_column()
