import openpyxl
from openpyxl.styles import Font
import os
import re
from zipfile import ZipFile

excel_dir_path = 'triage_result'
current_dir = os.path.abspath(os.path.dirname(__file__))
excel_abs_dir_path = os.path.join(current_dir, excel_dir_path)


def get_excels():
    dirs = os.listdir(excel_abs_dir_path)
    excel_list = []
    for excel in dirs:
        excel_list.append(os.path.join(excel_abs_dir_path, excel))
    return excel_list


def distinct_log_id_from_excel(excel_list):
    for excel in excel_list:
        match = re.match(r'(.*)\/(.*)\.xlsx', excel, re.I | re.M)
        excel_name = match.group(2)
        dest_filename = "{}-distinct.xlsx".format(excel_name)
        wb1 = openpyxl.load_workbook(excel)
        ws1 = wb1.active
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        #ws2 = wb2.create_sheet("error_log")
        values = []
        column_values = {}
        for i in range(2, ws1.max_row+1):
            if ws1.cell(row=i, column=9).value in column_values.keys():
                column_values[ws1.cell(row=i, column=9).value] += 1
            else:
                column_values[ws1.cell(row=i, column=9).value] = 1
                values.append(list(ws1[i]))

        column_names = ['', 'hostname', 'str_time', 'timestamp', 'id', 'msg', 'code_file', 'code_line',
                        'log_id', 'component', 'sub_component', 'argument_0', 'argument_1', 'level', 'log_id_count']
        ws2.append(column_names)
        row1 = ws2[1]
        for item in row1:
            item.font = Font(bold=True, size=11)

        for value in values:
            for key in column_values.keys():
                if value[8].value == key:
                    value.append(column_values[key])
                    break

        for value in values:
            value_list = []
            for item in value[0:14]:
                value_list.append(item.value)
            value_list.append(value[-1])
            print value_list
            ws2.append(value_list)

        wb2.save(filename=dest_filename)


def zip_excel():
    excel_file_list = get_excels()
    with ZipFile('triage_result/excels.zip', 'w') as zip:
        # writing each file one by one
        for file in excel_file_list:
            zip.write(file)

    print('All files zipped successfully!')


if __name__ == '__main__':
    excel_list = get_excels()
    print excel_list
    distinct_log_id_from_excel(excel_list)
    zip_excel()
