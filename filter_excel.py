import openpyxl
import os
import re
import argparse
import sys
import datetime
import calendar

excel_dir_path = 'triage_result'
current_dir = os.path.abspath(os.path.dirname(__file__))
excel_abs_dir_path = os.path.join(current_dir, excel_dir_path)


def get_excels(log_number):
    dirs = os.listdir(excel_abs_dir_path)
    for excel in dirs:
        if excel == "{}.xlsx".format(log_number):
            return os.path.join(excel_abs_dir_path, excel)


def filter_from_excel(excel, keyword_dict=None):
    match = re.match(r'(.*)\/(.*)\.xlsx', excel, re.I | re.M)
    excel_name = match.group(2)
    dest_filename = "{}-filter.xlsx".format(excel_name)
    wb1 = openpyxl.load_workbook(excel)
    ws1 = wb1.active
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    values = []
    if keyword_dict:
        for key, value in keyword_dict.items():
            if key == "none_logic":
                for i in range(2, ws1.max_row+1):
                    for v in value:
                        if str(ws1.cell(row=i, column=6).value).__contains__(v):
                            values.append(ws1[i])
                        else:
                            pass
            elif key == "OR":
                for i in range(2, ws1.max_row+1):
                    for v in value:
                        if str(ws1.cell(row=i, column=6).value).__contains__(v):
                            values.append(ws1[i])
                        else:
                            pass
            elif key == "AND":
                for i in range(2, ws1.max_row+1):
                    flag = True
                    for v in value:
                        if str(ws1.cell(row=i, column=6).value).__contains__(v):
                            flag = flag & True
                        else:
                            flag = flag & False
                    if flag:
                        values.append(ws1[i])
                    else:
                        pass

    column_names = ['', 'hostname', 'str_time', 'timestamp', 'syslog_identifier', 'msg', 'code_file', 'code_line',
                    'log_id', 'component', 'sub_component', 'argument_0', 'argument_1', 'level_name']
    ws2.append(column_names)

    for value in list(set(values)):
        value_list = []
        for item in value:
            value_list.append(item.value)
        print value_list
        ws2.append(value_list)

    wb2.save(filename=dest_filename)


def filter_from_excel_since_and_until(excel, since=None, until=None):
    match = re.match(r'(.*)\/(.*)\.xlsx', excel, re.I | re.M)
    excel_name = match.group(2)
    dest_filename = "{}-filter.xlsx".format(excel_name)
    wb1 = openpyxl.load_workbook(excel)
    ws1 = wb1.active
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    values = []

    for i in range(2, ws1.max_row+1):
        if since <= ws1.cell(row=i, column=4).value <= until:
            values.append(ws1[i])
        else:
            pass

    column_names = ['', 'hostname', 'str_time', 'timestamp', 'syslog_identifier', 'msg', 'code_file', 'code_line',
                    'log_id', 'component', 'sub_component', 'argument_0', 'argument_1', 'level_name']
    ws2.append(column_names)

    for value in list(set(values)):
        value_list = []
        for item in value:
            value_list.append(item.value)
        print value_list
        ws2.append(value_list)

    wb2.save(filename=dest_filename)


def filter_from_excel_with_msg_since_and_until(excel, keyword_dict=None, since=None, until=None):
    match = re.match(r'(.*)\/(.*)\.xlsx', excel, re.I | re.M)
    excel_name = match.group(2)
    dest_filename = "{}-filter.xlsx".format(excel_name)
    wb1 = openpyxl.load_workbook(excel)
    ws1 = wb1.active
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    values = []
    if keyword_dict:
        for key, value in keyword_dict.items():
            if key == "none_logic":
                for i in range(2, ws1.max_row+1):
                    for v in value:
                        if str(ws1.cell(row=i, column=6).value).__contains__(v) and since <= ws1.cell(row=i, column=4).value <= until:
                            values.append(ws1[i])
                        else:
                            pass
            elif key == "OR":
                for i in range(2, ws1.max_row+1):
                    for v in value:
                        if str(ws1.cell(row=i, column=6).value).__contains__(v) and since <= ws1.cell(row=i, column=4).value <= until:
                            values.append(ws1[i])
                        else:
                            pass
            elif key == "AND":
                for i in range(2, ws1.max_row+1):
                    flag = True
                    for v in value:
                        if str(ws1.cell(row=i, column=6).value).__contains__(v):
                            flag = flag & True
                        else:
                            flag = flag & False
                    if flag and since <= ws1.cell(row=i, column=4).value <= until:
                        values.append(ws1[i])
                    else:
                        pass

    column_names = ['', 'hostname', 'str_time', 'timestamp', 'syslog_identifier', 'msg', 'code_file', 'code_line',
                    'log_id', 'component', 'sub_component', 'argument_0', 'argument_1', 'level_name']
    ws2.append(column_names)

    for value in list(set(values)):
        value_list = []
        for item in value:
            value_list.append(item.value)
        print value_list
        ws2.append(value_list)

    wb2.save(filename=dest_filename)


def parse_word(keyword_list):
    keyword_dict = {}
    keyword_list = set(keyword_list)
    or_list = ['OR', 'Or', 'or', 'oR']
    and_list = ['AND', 'and', 'And', 'ANd', 'aND', 'anD', 'aNd', 'AnD']
    for item in or_list:
        if keyword_list.__contains__(item):
            keyword_list.remove(item)
            keyword_dict["OR"] = keyword_list
            break
    for item in and_list:
        if keyword_list.__contains__(item):
            keyword_list.remove(item)
            keyword_dict["AND"] = keyword_list
            break
    if keyword_dict == {}:
        keyword_dict["none_logic"] = keyword_list
    else:
        pass
    return keyword_dict


def parse_time_string(start_time_string, end_time_string):

    start_timestamp = calendar.timegm(datetime.datetime.strptime(start_time_string, "%Y-%m-%d %H:%M:%S").timetuple())
    end_timestamp = calendar.timegm(datetime.datetime.strptime(end_time_string, "%Y-%m-%d %H:%M:%S").timetuple())

    return start_timestamp * 1000000, end_timestamp * 1000000


def main():
    parser = argparse.ArgumentParser(description="Excel filter tool.")
    parser.add_argument('-path', '-p', dest="p", action="append", default=None,
                        help='data path.')
    parser.add_argument('-msg', '-m', dest="m", action='append', default=None,
                        help='Key message word to dc.')
    parser.add_argument('-since', '-s', dest="s", action="store", default=None,
                        help='Start the searching from the specified date time.')
    parser.add_argument('-until', '-u', dest="u", action="store", default=None,
                        help='End the searching from the specified date time.')

    try:
        global args
        args = parser.parse_args()
    except:
        parser.print_help()
        sys.exit(1)

    if args.p is None and args.m is None:
        pass
    elif args.p is None and (args.s is None or args.u is None):
        pass
    elif args.p and args.s and args.u and args.m:
        for path in args.p:
            match = re.match('(.*)\/(log.*)', path, re.I | re.M)
            log_number = match.group(2)
            excel = get_excels(log_number)
            print excel
            keyword = args.m
            keyword_dict = parse_word(keyword)
            since = args.s
            until = args.u
            since_timestamp, until_timestamp = parse_time_string(since, until)
            filter_from_excel_with_msg_since_and_until(excel, keyword_dict, since_timestamp, until_timestamp)
    elif args.p and args.m:
        for path in args.p:
            match = re.match('(.*)\/(log.*)', path, re.I | re.M)
            log_number = match.group(2)
            excel = get_excels(log_number)
            print excel
            keyword = args.m
            keyword_dict = parse_word(keyword)
            filter_from_excel(excel, keyword_dict)
    elif args.p and args.s and args.u:
        for path in args.p:
            match = re.match('(.*)\/(log.*)', path, re.I | re.M)
            log_number = match.group(2)
            excel = get_excels(log_number)
            print excel
            since = args.s
            until = args.u
            since_timestamp, until_timestamp = parse_time_string(since, until)
            filter_from_excel_since_and_until(excel, since_timestamp, until_timestamp)


if __name__ == '__main__':
    main()
